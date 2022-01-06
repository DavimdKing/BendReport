
from Script.SortingandFiltering import Total_bend_rate_per_month_mcc, Total_bend_rate_per_month, \
    Total_bend_rate_per_week, SO_bend_rate_per_month, SO_bend_rate_per_week, MPT_bend_rate_per_month, \
    MPT_bend_rate_per_week, HTO_bend_rate_per_month, HTO_bend_rate_per_week, CDM_MPT_month, CDM_HTO_month, \
    CDM_SO_month, CDM_SO_week, CDM_MPT_week, CDM_HTO_week, CDM_Total_week, main_df, SO_permonth, SO_perweek, \
    MPT_permonth, MPT_perweek, HTO_permonth, HTO_perweek, CDM_numberoflot_week, CDM_numberoflot_month
import pandas as pd
from pandas import DataFrame
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from Script.TemplateForWeeklyBendReport import bend_report_template, month_sorting_order, week_sorting_order, \
    new_week_sorting_order, new_month_sorting_order, template_date_list, sorting_month_template, sorting_week_template, \
    lot_counting_template, lot_number_template
from Script.Lot_counting import total_week_lot_list, total_week_lot_list_mmc, Lot_counting

Constant_a = []

wb = Workbook()
ws = wb.active
ws.title = "Week"

bend_report_template(ws)

for i in range(len(Total_bend_rate_per_month)):
    ws[get_column_letter(i + 3) + str(11)] = \
        (Total_bend_rate_per_month.loc[:, 'bend_rate_per_month'])[i]
    ws[get_column_letter(i + 3) + str(40)] = \
        (Total_bend_rate_per_month.loc[:, 'bend_rate_per_month'])[i]

for i in range(len(SO_bend_rate_per_month)):
    ws[get_column_letter(i + 3 + sorting_month_template(SO_permonth)[i]) + str(41)] = \
        (SO_bend_rate_per_month.loc[:, 'bend_rate_per_month'])[i]

for i in range(len(MPT_bend_rate_per_month)):
    ws[get_column_letter(i + 3 + sorting_month_template(MPT_permonth)[i]) + str(42)] = \
        (MPT_bend_rate_per_month.loc[:, 'bend_rate_per_month'])[i]

for i in range(len(HTO_bend_rate_per_month)):
    ws[get_column_letter(i + 3 + sorting_month_template(HTO_permonth)[i]) + str(43)] = \
        (HTO_bend_rate_per_month.loc[:, 'bend_rate_per_month'])[i]

Total_bend_rate_per_month_mcc = Total_bend_rate_per_month_mcc.sort_values((["CDM机号", 'Month_number']))
for i in range(5):
    Constant_a = Total_bend_rate_per_month_mcc.iloc[(int(len(Total_bend_rate_per_month_mcc) / 5) * (i)) \
                                                    :(int(len(Total_bend_rate_per_month_mcc) / 5) * (1 + i)), :]
    for j in range(int(len(Total_bend_rate_per_month_mcc) / 5)):
        ws[get_column_letter(j + 3) + str(6 + i)] = \
            (list(Constant_a.loc[:, 'bend_rate_per_month']))[j]
        ws[get_column_letter(j + 3) + str(15 + i * 5)] = \
            (list(Constant_a.loc[:, 'bend_rate_per_month']))[j]

Constant_b = [11, 40]
for j in range(len(Constant_b)):
    for i in range(len(Total_bend_rate_per_week)):
        ws[get_column_letter(4 + len(template_date_list(main_df)) + i) + str(f'{Constant_b[j]}')] = \
            (list(Total_bend_rate_per_week.loc[:, 'bend_rate_per_week']))[i]

# for i in range(len(Total_bend_rate_per_week)):
#     ws[get_column_letter(6+new_week_sorting_order(Total_bend_rate_per_week)[i]) + str(40)] = \
#         (list(Total_bend_rate_per_week.loc[:, 'bend_rate_per_week']))[i]

# try:
#     for i in range(len(CDM_SO_month['CDM1'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_SO_month['CDM1'])) + str(16)] = \
#             (list(CDM_SO_month['CDM1'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_SO_month['CDM2'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_SO_month['CDM2'])) + str(21)] = \
#             (list(CDM_SO_month['CDM2'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_SO_month['CDM3'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_SO_month['CDM3'])) + str(26)] = \
#             (list(CDM_SO_month['CDM3'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_SO_month['CDM4'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_SO_month['CDM4'])) + str(31)] = \
#             (list(CDM_SO_month['CDM4'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_SO_month['CDM5'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_SO_month['CDM5'])) + str(36)] = \
#             (list(CDM_SO_month['CDM5'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
#
# try:
#     for i in range(len(CDM_MPT_month['CDM1'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_MPT_month['CDM1'])) + str(17)] = \
#             (list(CDM_MPT_month['CDM1'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_MPT_month['CDM2'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_MPT_month['CDM2'])) + str(22)] = \
#             (list(CDM_MPT_month['CDM2'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_MPT_month['CDM3'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_MPT_month['CDM3'])) + str(27)] = \
#             (list(CDM_MPT_month['CDM3'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_MPT_month['CDM4'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_MPT_month['CDM4'])) + str(32)] = \
#             (list(CDM_MPT_month['CDM4'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_MPT_month['CDM5'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_MPT_month['CDM5'])) + str(37)] = \
#             (list(CDM_MPT_month['CDM5'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass

# try:
#     for i in range(len(CDM_HTO_month['CDM1'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_HTO_month['CDM1'])) + str(18)] = \
#             (list(CDM_HTO_month['CDM1'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_HTO_month['CDM2'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_HTO_month['CDM2'])) + str(23)] = \
#             (list(CDM_HTO_month['CDM2'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_HTO_month['CDM3'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_HTO_month['CDM3'])) + str(28)] = \
#             (list(CDM_HTO_month['CDM3'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_HTO_month['CDM4'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_HTO_month['CDM4'])) + str(33)] = \
#             (list(CDM_HTO_month['CDM4'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass
# try:
#     for i in range(len(CDM_HTO_month['CDM5'])):
#         ws[get_column_letter(i + 2 + month_sorting_order(CDM_HTO_month['CDM5'])) + str(38)] = \
#             (list(CDM_HTO_month['CDM5'].loc[:, 'bend_rate_per_month']))[i]
# except KeyError:
#     pass

# for i in range(len(CDM_SO_week['CDM2'])):
#     ws[get_column_letter(i + 15 + week_sorting_order(CDM_SO_week['CDM2'])) + str(21)] = \
#         (list(CDM_SO_week['CDM2'].loc[:, 'bend_rate_per_week']))[i]

# for i in range(1, 6):
#     try:
#         for x in range(len(CDM_SO_week[f'CDM{i}'])):
#             ws[get_column_letter(x + 15 + week_sorting_order(CDM_SO_week[f'CDM{i}'])) + str(16+5*(i-1))] = \
#                 (list(CDM_SO_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
#     except KeyError:
#         pass
# for i in range(1, 6):
#     try:
#         for x in range(len(CDM_MPT_week[f'CDM{i}'])):
#             ws[get_column_letter(x + 15 + week_sorting_order(CDM_MPT_week[f'CDM{i}'])) + str(17+5*(i-1))] = \
#                 (list(CDM_MPT_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
#     except KeyError:
#         pass
# for i in range(1, 6):
#     try:
#         for x in range(len(CDM_HTO_week[f'CDM{i}'])):
#             ws[get_column_letter(x + 15 + week_sorting_order(CDM_HTO_week[f'CDM{i}'])) + str(18+5*(i-1))] = \
#                 (list(CDM_HTO_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
#     except KeyError:
#         pass

for i in range(len(SO_bend_rate_per_week)):
    ws[get_column_letter(4 + len(template_date_list(main_df)) + i + sorting_week_template(SO_perweek)[i]) + str(41)] = \
        (list(SO_bend_rate_per_week.loc[:, 'bend_rate_per_week']))[i]

for i in range(len(MPT_bend_rate_per_week)):
    ws[get_column_letter(4 + len(template_date_list(main_df)) + i + sorting_week_template(MPT_perweek)[i]) + str(42)] = \
        (list(MPT_bend_rate_per_week.loc[:, 'bend_rate_per_week']))[i]

for i in range(len(HTO_bend_rate_per_week)):
    ws[get_column_letter(4 + len(template_date_list(main_df)) + i + sorting_week_template(HTO_perweek)[i]) + str(43)] = \
        (list(HTO_bend_rate_per_week.loc[:, 'bend_rate_per_week']))[i]

for i in range(1, 6):
    try:
        for x in range(len(CDM_Total_week[f'CDM{i}'])):
            ws[get_column_letter(4 + len(template_date_list(main_df)) + x) + str(15 + 5 * (i - 1))] = \
                (list(CDM_Total_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
            ws[get_column_letter(4 + len(template_date_list(main_df)) + x) + str(6 + 1 * (i - 1))] = \
                (list(CDM_Total_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
    except KeyError:
        pass

for i in range(1, 6):
    try:
        for x in range(len(CDM_HTO_week[f'CDM{i}'])):
            ws[get_column_letter(4 + len(template_date_list(main_df)) + x + sorting_week_template(HTO_perweek)[x]) + \
               str(18 + 5 * (i - 1))] = (list(CDM_HTO_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
    except KeyError:
        pass
for i in range(1, 6):
    try:
        for x in range(len(CDM_HTO_month[f'CDM{i}'])):
            ws[get_column_letter(x + 3 + sorting_month_template(HTO_permonth)[x]) + str(18 + 5 * (i - 1))] = \
                (list(CDM_HTO_month[f'CDM{i}'].loc[:, 'bend_rate_per_month']))[x]
    except KeyError:
        pass
for i in range(1, 6):
    try:
        for x in range(len(CDM_MPT_week[f'CDM{i}'])):
            ws[get_column_letter(4 + len(template_date_list(main_df)) + x + sorting_week_template(MPT_perweek)[x]) + \
               str(17 + 5 * (i - 1))] = (list(CDM_MPT_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
    except KeyError:
        pass
for i in range(1, 6):
    try:
        for x in range(len(CDM_MPT_month[f'CDM{i}'])):
            ws[get_column_letter(x + 3 + sorting_month_template(MPT_permonth)[x]) + str(17 + 5 * (i - 1))] = \
                (list(CDM_MPT_month[f'CDM{i}'].loc[:, 'bend_rate_per_month']))[x]
    except KeyError:
        pass
for i in range(1, 6):
    try:
        for x in range(len(CDM_SO_week[f'CDM{i}'])):
            ws[get_column_letter(4 + len(template_date_list(main_df)) + x + sorting_week_template(SO_perweek)[x]) + \
               str(16 + 5 * (i - 1))] = (list(CDM_SO_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
    except KeyError:
        pass
for i in range(1, 6):
    try:
        for x in range(len(CDM_SO_month[f'CDM{i}'])):
            ws[get_column_letter(x + 3 + sorting_month_template(SO_permonth)[x]) + str(16 + 5 * (i - 1))] = \
                (list(CDM_SO_month[f'CDM{i}'].loc[:, 'bend_rate_per_month']))[x]
    except KeyError:
        pass

for row in range(6):
    for cell in range(len(Total_bend_rate_per_month)):
        (ws[get_column_letter(cell + 3) + str(6 + row)]) \
            .number_format = "0.00%"
    for cell in range(len(Total_bend_rate_per_week)):
        (ws[get_column_letter(4 + len(template_date_list(main_df)) + cell) + str(6 + row)]) \
            .number_format = "0.00%"
for row in range(29):
    for cell in range(len(Total_bend_rate_per_month)):
        (ws[get_column_letter(cell + 3) + str(15 + row)]) \
            .number_format = "0.00%"
    for cell in range(len(Total_bend_rate_per_week)):
        (ws[get_column_letter(4 + len(template_date_list(main_df)) + cell) + str(15 + row)]) \
            .number_format = "0.00%"

new_worksheet = wb.create_sheet('lot_counting')
ws2 = wb['lot_counting']
lot_counting_template(ws2)

for i in range(3):
    SO_value, SO_machine = Lot_counting('SO')
    HTO_value, HTO_machine = Lot_counting('HTO')
    MPT_value, MPT_machine = Lot_counting('MPT')
    ws2['C' + str(4 + i)] = len(total_week_lot_list[i])
    ws2['E' + str(4 + i)] = SO_value[i]
    ws2['G' + str(4 + i)] = MPT_value[i]
    ws2['I' + str(4 + i)] = HTO_value[i]
    for x in range(5):
        ws2['C' + str(10 + x * 6 + 1 * i)] = len(total_week_lot_list_mmc[5 * i + x])
        ws2['E' + str(10 + x * 6 + 1 * i)] = SO_machine[5 * i + x]
        ws2['G' + str(10 + x * 6 + 1 * i)] = MPT_machine[5 * i + x]
        ws2['I' + str(10 + x * 6 + 1 * i)] = HTO_machine[5 * i + x]

for i in range(1, 6):
    try:
        for x in range(len(CDM_Total_week[f'CDM{i}'][-3:])):
            ws2['D' + str(12 - x * 1 + (i - 1) * 6)] = \
                (CDM_Total_week[f'CDM{i}'][-3:].iloc[x, -1])
    except KeyError:
        pass
Three_weeks = sorted(list(main_df['Week_number'].drop_duplicates()))[-3:]
for i in range(1, 6):
    for x in range(len(Three_weeks)):
        try:
            ws2['J' + str(12 + (i - 1) * 6 - 1 * x)] = (CDM_HTO_week[f'CDM{i}'][CDM_HTO_week[f'CDM{i}'].Week_number == \
                                                                            Three_weeks[x]].iloc[0,-1])
        except IndexError:
            pass
        except KeyError:
            pass
    for x in range(len(Three_weeks)):
        try:
            ws2['F' + str(12 + (i - 1) * 6 - 1 * x)] = (CDM_SO_week[f'CDM{i}'][CDM_SO_week[f'CDM{i}'].Week_number == \
                                                                            Three_weeks[x]].iloc[0,-1])
        except IndexError:
            pass
        except KeyError:
            pass
    for x in range(len(Three_weeks)):
        try:
            ws2['H' + str(12 + (i - 1) * 6 - 1 * x)] = (CDM_MPT_week[f'CDM{i}'][CDM_MPT_week[f'CDM{i}'].Week_number == \
                                                                            Three_weeks[x]].iloc[0,-1])
        except IndexError:
            pass
        except KeyError:
            pass

for x in range(len(Three_weeks)):
    try:
        ws2['J' + str(6 - 1 * x)] = (HTO_bend_rate_per_week[HTO_bend_rate_per_week.Week_number ==\
                                                            Three_weeks[x]].iloc[0,-1])
    except IndexError:
        pass
    except KeyError:
        pass
for x in range(len(Three_weeks)):
    try:
        ws2['F' + str(6 - 1 * x)] = (SO_bend_rate_per_week[SO_bend_rate_per_week.Week_number ==\
                                                            Three_weeks[x]].iloc[0,-1])
    except IndexError:
        pass
    except KeyError:
        pass
for x in range(len(Three_weeks)):
    try:
        ws2['H' + str(6 - 1 * x)] = (MPT_bend_rate_per_week[MPT_bend_rate_per_week.Week_number ==\
                                                            Three_weeks[x]].iloc[0,-1])
    except IndexError:
        pass
    except KeyError:
        pass
for x in range(len(Three_weeks)):
    try:
        ws2['D' + str(6 - 1 * x)] = (Total_bend_rate_per_week[Total_bend_rate_per_week.Week_number ==\
                                                            Three_weeks[x]].iloc[0,-1])
    except IndexError:
        pass
    except KeyError:
        pass


# (CDM_Total_week['CDM2'][-3:].iloc[2, -1])
#
# for i in range (len(All_HTO_lots)):
#     ws['C' + str(i + 2)] = All_HTO_lots[i]
#
# for i in range (len(All_SO_lots)):
#     ws['D' + str(i + 2)] = All_SO_lots[i]
#
# for i in range (len(All_Other_lots)):
#     ws['E' + str(i + 2)] = All_Other_lots[i]
#

new_worksheet = wb.create_sheet('lot_number')
ws3 = wb['lot_number']
lot_number_template(ws3)
for x in range(1, 6):
    for i in range(Total_bend_rate_per_week.shape[0]):
        ws3[get_column_letter(11+i) + str(5+x)] = (CDM_numberoflot_week.groupby(CDM_numberoflot_week.CDM机号).get_group(x).\
            reset_index(drop=True)['生产批号'][i])
    for j in range(Total_bend_rate_per_month.shape[0]):
        ws3[get_column_letter(3 + j) + str(5 + x)] = (CDM_numberoflot_month.groupby(CDM_numberoflot_month.CDM机号).\
            get_group(x).reset_index(drop=True)['生产批号'][j])


wb.save('Weektestsummary.xlsx')
