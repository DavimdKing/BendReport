from tkinter import *
from tkinter import filedialog

def excel_input(defect, address):
    from Bin.All2021_Weeks import All_Weeks
    from pandas import Series
    from pandas import DataFrame
    import pandas as pd
    from collections import Counter
    from Script.Function_for_sorting_data import floating_to_percentage_week, floating_to_percentage_month

    Type_of_defect = defect  # bend = 'Bend'& contamination = 'Co + Adhesive Co'
    File_address = address
    # for example, '//192.168.80.20/share/Bu1-Eng/David/Bend/Bend summary2.xlsx'

    All_MPT_lots = []
    All_HTO_lots = []
    All_SO_lots = []
    All_Other_lots = []
    Number_of_MPT = 0
    Number_of_HTO = 0
    Number_of_SO = 0
    Number_of_Others = 0
    lots_number = ()
    CDM_Date = []
    CDM_machine_no = []
    MPT_perweek = []
    MPT_permonth = []
    SO_perweek = []
    SO_permonth = []
    HTO_perweek = []
    HTO_permonth = []
    Others_perweek = []
    Others_permonth = []
    Amount_of_Lotperweek = []
    Amount_of_Lotpermonth = []

    MPT_program = ('L-16420-08N04Z', 'L-16460-07N04Z', 'L-17290-06N04Z', 'L-17300-05N04Z', 'L-17310-06N04Z',
                   'L-17370-04N04Z', 'L-17580-07N04Z', 'L-17720-02N04Z', 'L-17860-00N04Z', 'L-15750-06N04Z',
                   'H-4700-04N04Z', 'L-15670-07N04Z', 'L-17460-04N04Z', 'L-15880-04 TAB.AN04Z', 'L-15880-04 TAB.BN04Z',
                   'L-17480-04N04Z', 'L-17470-04N04Z', 'L-18290-02N04Z', 'L-18620-01N04Z', 'L-19120-02N04Z',
                   'L-19010-01N04Z', 'L-19300-01N04Z', 'L-17460-04N04', 'L-17480-04N04', 'L-17480-05N04',
                   'L-18290-02N04',
                   'L-18620-01N04', 'L-19010-01N04', 'L-19120-02N04', 'L-19300-01N04', 'H-4700-04N04', 'L-17370-06N04',
                   'L-17370-04N04', 'L-19320-01N04', 'L-19430-01 OPT1', 'L-19320-00N04', 'L-15750-06N04',
                   'L-17480-05N04Z'
                   )
    HTO_program = {'L-18250-02N04Z', 'L-19110-03N04Z', 'L-18040-01N04Z', 'L-18600-00N04Z', 'L-19500-01N04Z',
                   'L-18250-02N04', 'L-18530-02N04Z'
                   }
    SO_program = {'SL-13340-EN04', 'SL-13790-BN04', 'SL-13800-BN04', 'SL-13590-BN04', 'SL-13850-CN04', 'SL-13870-AN04', \
                  'SL-13900-AN04'
                  }

    main_df = pd.read_excel(File_address, sheet_name='8', )
    main_df = main_df.rename({'产品\n名称': '产品名称', '生产\n批号': '生产批号', 'CDM生产\n时间': 'CDM生产时间', 'CDM\n机号': \
        'CDM机号', '产品\n总个数': '产品总个数'}, axis=1)

    lots_program = main_df['产品名称']
    lots_number = main_df['生产批号']
    CDM_Date = main_df['CDM生产时间']
    CDM_machine_no = main_df['CDM机号']
    Total_product_number = main_df['产品总个数']
    Total_bend_number = main_df[Type_of_defect]
    All_bend_rate = main_df['rate[%]']

    for i in lots_program:
        if i in MPT_program:

            Number_of_MPT = Number_of_MPT + 1
        elif i in HTO_program:

            Number_of_HTO = Number_of_HTO + 1
        elif i in SO_program:

            Number_of_SO = Number_of_SO + 1
        else:

            Number_of_Others = Number_of_Others + 1

    index_of_MPT_lots = []
    index_of_HTO_lots = []
    index_of_SO_lots = []
    index_of_Other_lots = []
    All_MPT_Date = []
    All_SO_Date = []
    All_HTO_Date = []
    All_Other_Date = []

    for i in MPT_program:
        index_of_MPT_lots.extend(lots_program[lots_program == i].index)
    All_MPT_lots = lots_number[index_of_MPT_lots].values

    All_MPT_Date = CDM_Date[index_of_MPT_lots].values
    MPT_df = main_df.iloc[index_of_MPT_lots, :].reset_index()
    main_df['customer'] = ''
    main_df.iloc[index_of_MPT_lots, -1] = 'MPT'

    for i in HTO_program:
        index_of_HTO_lots.extend(lots_program[lots_program == i].index)
    All_HTO_lots = lots_number[index_of_HTO_lots].values
    All_HTO_Date = CDM_Date[index_of_HTO_lots].values
    HTO_df = main_df.iloc[index_of_HTO_lots, :].reset_index()
    main_df.iloc[index_of_HTO_lots, -1] = 'HTO'

    for i in SO_program:
        index_of_SO_lots.extend(lots_program[lots_program == i].index)
    All_SO_lots = lots_number[index_of_SO_lots].values
    All_SO_Date = CDM_Date[index_of_SO_lots].values
    SO_df = main_df.iloc[index_of_SO_lots, :].reset_index()
    main_df.iloc[index_of_SO_lots, -1] = 'SO'

    All_Other_lots = list(
        ((set(lots_number).difference(set(All_MPT_lots))).difference((set(All_HTO_lots)))).difference(
            (set(All_SO_lots))))

    for i in All_Other_lots:
        index_of_Other_lots.extend(lots_number[lots_number == i].index)
    All_Other_Date = CDM_Date[index_of_Other_lots].values
    Other_df = main_df.iloc[index_of_Other_lots, :].reset_index()

    for i in All_MPT_Date:
        MPT_perweek.append(All_Weeks.week_searching(pd.Timestamp(i)))
        MPT_permonth.append(All_Weeks.month_searching(i))
    # print(Counter(MPT_perweek).items())

    # CDM_perweek = []
    # CDM_permonth = []
    # for i in CDM_Date:
    #     CDM_perweek.append(All_Weeks.week_searching(pd.Timestamp(i)))
    #     CDM_permonth.append(All_Weeks.month_searching(i))
    # print(CDM_permonth)

    for i in All_SO_Date:
        SO_perweek.append(All_Weeks.week_searching(pd.Timestamp(i)))
        SO_permonth.append(All_Weeks.month_searching(i))

    for i in All_HTO_Date:
        HTO_perweek.append(All_Weeks.week_searching(pd.Timestamp(i)))
        HTO_permonth.append(All_Weeks.month_searching(i))
    for i in All_Other_Date:
        Others_perweek.append(All_Weeks.week_searching(pd.Timestamp(i)))
        Others_permonth.append(All_Weeks.month_searching(i))

    for i in CDM_Date:
        Amount_of_Lotperweek.append(All_Weeks.week_searching(i))
        Amount_of_Lotpermonth.append(All_Weeks.month_searching(i))
    # print(Counter(Amount_of_Lotperweek).items())

    main_df['Week_number'] = pd.DataFrame(Amount_of_Lotperweek)
    main_df['Month_number'] = pd.DataFrame(Amount_of_Lotpermonth)
    Total_bend_rate_per_week = main_df.groupby('Week_number')[['产品总个数', Type_of_defect]].sum()
    Total_bend_rate_per_week['bend_rate_per_week'] = (
            Total_bend_rate_per_week.iloc[:, 1] / Total_bend_rate_per_week.iloc[:, 0])
    Total_bend_rate_per_week = Total_bend_rate_per_week.reset_index()
    # Total_bend_rate_per_week['bend_rate_per_week'] = pd.Series(["{0:.2f}%".format(val * 100) for val in \
    #     Total_bend_rate_per_week['bend_rate_per_week']], index = Total_bend_rate_per_week.index)
    floating_to_percentage_week((Total_bend_rate_per_week))

    Total_bend_rate_per_week_mcc = main_df.groupby(['Week_number', 'CDM机号'])[['产品总个数', Type_of_defect]].sum().unstack(
        fill_value=0).stack()
    Total_bend_rate_per_week_mcc['bend_rate_per_week'] = (
            Total_bend_rate_per_week_mcc.iloc[:, 1] / Total_bend_rate_per_week_mcc.iloc[:, 0])
    Total_bend_rate_per_week_mcc = Total_bend_rate_per_week_mcc.reset_index()
    Total_bend_rate_per_week_mcc.iloc[:, -1] = Total_bend_rate_per_week_mcc.fillna(0).iloc[:, -1]
    # Total_bend_rate_per_week_mcc['bend_rate_per_week'] = pd.Series(["{0:.2f}%".format(val * 100) for val in \
    #     Total_bend_rate_per_week_mcc['bend_rate_per_week']], index = Total_bend_rate_per_week_mcc.index)
    floating_to_percentage_week(Total_bend_rate_per_week_mcc)

    Total_bend_rate_per_month = main_df.groupby('Month_number')[['产品总个数', Type_of_defect]].sum()
    Total_bend_rate_per_month['bend_rate_per_month'] = (
            Total_bend_rate_per_month.iloc[:, 1] / Total_bend_rate_per_month.iloc[:, 0])
    Total_bend_rate_per_month = Total_bend_rate_per_month.reset_index()
    # Total_bend_rate_per_month['bend_rate_per_month'] = pd.Series(["{0:.2f}%".format(val * 100) for val in \
    #     Total_bend_rate_per_month['bend_rate_per_month']], index = Total_bend_rate_per_month.index)
    floating_to_percentage_month(Total_bend_rate_per_month)

    Total_bend_rate_per_month_mcc = main_df.groupby(['Month_number', 'CDM机号'])[['产品总个数', Type_of_defect]].sum().unstack(
        fill_value=0).stack()
    Total_bend_rate_per_month_mcc['bend_rate_per_month'] = (
            Total_bend_rate_per_month_mcc.iloc[:, 1] / Total_bend_rate_per_month_mcc.iloc[:, 0])
    Total_bend_rate_per_month_mcc = Total_bend_rate_per_month_mcc.reset_index()
    Total_bend_rate_per_month_mcc.iloc[:, -1] = Total_bend_rate_per_month_mcc.fillna(0).iloc[:, -1]
    # Total_bend_rate_per_month_mcc['bend_rate_per_month'] = pd.Series(["{0:.2f}%".format(val * 100) for val in \
    #     Total_bend_rate_per_month_mcc['bend_rate_per_month']], index = Total_bend_rate_per_month_mcc.index)
    floating_to_percentage_month(Total_bend_rate_per_month_mcc)

    MPT_df['Week_number'] = pd.DataFrame(MPT_perweek)
    MPT_df['Month_number'] = pd.DataFrame(MPT_permonth)
    MPT_bend_rate_per_week = MPT_df.groupby('Week_number')[['产品总个数', Type_of_defect]].sum()
    MPT_bend_rate_per_week['bend_rate_per_week'] = (
                MPT_bend_rate_per_week.iloc[:, 1] / MPT_bend_rate_per_week.iloc[:, 0])
    MPT_bend_rate_per_week = MPT_bend_rate_per_week.reset_index()
    # MPT_bend_rate_per_week['bend_rate_per_week'] = pd.Series(["{0:.2f}%".format(val * 100) for val in \
    #     MPT_bend_rate_per_week['bend_rate_per_week']], index = MPT_bend_rate_per_week.index)
    floating_to_percentage_week(MPT_bend_rate_per_week)

    MPT_bend_rate_per_week_mcc = MPT_df.groupby(['Week_number', 'CDM机号'])[['产品总个数', Type_of_defect]].sum().unstack(
        fill_value=0).stack()
    MPT_bend_rate_per_week_mcc['bend_rate_per_week'] = (
            MPT_bend_rate_per_week_mcc.iloc[:, 1] / MPT_bend_rate_per_week_mcc.iloc[:, 0])
    MPT_bend_rate_per_week_mcc = MPT_bend_rate_per_week_mcc.reset_index()
    MPT_bend_rate_per_week_mcc.iloc[:, -1] = MPT_bend_rate_per_week_mcc.fillna(0).iloc[:, -1]
    floating_to_percentage_week(MPT_bend_rate_per_week_mcc)

    # MPT_bend_rate_per_week_mcc['bend_rate_per_week'] = pd.Series(["{0:.2f}%".format(val * 100) for val in \
    #     MPT_bend_rate_per_week_mcc['bend_rate_per_week']], index = MPT_bend_rate_per_week_mcc.index)

    MPT_bend_rate_per_month = MPT_df.groupby('Month_number')[['产品总个数', Type_of_defect]].sum()
    MPT_bend_rate_per_month['bend_rate_per_month'] = (
            MPT_bend_rate_per_month.iloc[:, 1] / MPT_bend_rate_per_month.iloc[:, 0])
    MPT_bend_rate_per_month = MPT_bend_rate_per_month.reset_index()
    floating_to_percentage_month(MPT_bend_rate_per_month)
    # MPT_bend_rate_per_month['bend_rate_per_month'] = pd.Series(["{0:.2f}%".format(val * 100) for val in \
    #     MPT_bend_rate_per_month['bend_rate_per_month']], index = MPT_bend_rate_per_month.index)

    MPT_bend_rate_per_month_mcc = MPT_df.groupby(['Month_number', 'CDM机号'])[['产品总个数', Type_of_defect]].sum().unstack(
        fill_value=0).stack()
    MPT_bend_rate_per_month_mcc['bend_rate_per_month'] = (
            MPT_bend_rate_per_month_mcc.iloc[:, 1] / MPT_bend_rate_per_month_mcc.iloc[:, 0])
    MPT_bend_rate_per_month_mcc = MPT_bend_rate_per_month_mcc.reset_index()
    MPT_bend_rate_per_month_mcc.iloc[:, -1] = MPT_bend_rate_per_month_mcc.fillna(0).iloc[:, -1]
    floating_to_percentage_month(MPT_bend_rate_per_month_mcc)

    HTO_df['Week_number'] = pd.DataFrame(HTO_perweek)
    HTO_df['Month_number'] = pd.DataFrame(HTO_permonth)
    HTO_bend_rate_per_week = HTO_df.groupby('Week_number')[['产品总个数', Type_of_defect]].sum()
    HTO_bend_rate_per_week['bend_rate_per_week'] = (
                HTO_bend_rate_per_week.iloc[:, 1] / HTO_bend_rate_per_week.iloc[:, 0])
    HTO_bend_rate_per_week = HTO_bend_rate_per_week.reset_index()
    floating_to_percentage_week(HTO_bend_rate_per_week)

    HTO_bend_rate_per_week_mcc = HTO_df.groupby(['Week_number', 'CDM机号'])[['产品总个数', Type_of_defect]].sum().unstack(
        fill_value=0).stack()
    HTO_bend_rate_per_week_mcc['bend_rate_per_week'] = (
            HTO_bend_rate_per_week_mcc.iloc[:, 1] / HTO_bend_rate_per_week_mcc.iloc[:, 0])
    HTO_bend_rate_per_week_mcc = HTO_bend_rate_per_week_mcc.reset_index()
    HTO_bend_rate_per_week_mcc.iloc[:, -1] = HTO_bend_rate_per_week_mcc.fillna(0).iloc[:, -1]
    floating_to_percentage_week(HTO_bend_rate_per_week_mcc)

    HTO_bend_rate_per_month = HTO_df.groupby('Month_number')[['产品总个数', Type_of_defect]].sum()
    HTO_bend_rate_per_month['bend_rate_per_month'] = (
            HTO_bend_rate_per_month.iloc[:, 1] / HTO_bend_rate_per_month.iloc[:, 0])
    HTO_bend_rate_per_month = HTO_bend_rate_per_month.reset_index()
    floating_to_percentage_month(HTO_bend_rate_per_month)

    HTO_bend_rate_per_month_mcc = HTO_df.groupby(['Month_number', 'CDM机号'])[['产品总个数', Type_of_defect]].sum().unstack(
        fill_value=0).stack()
    HTO_bend_rate_per_month_mcc['bend_rate_per_month'] = (
            HTO_bend_rate_per_month_mcc.iloc[:, 1] / HTO_bend_rate_per_month_mcc.iloc[:, 0])
    HTO_bend_rate_per_month_mcc = HTO_bend_rate_per_month_mcc.reset_index()
    HTO_bend_rate_per_month_mcc.iloc[:, -1] = HTO_bend_rate_per_month_mcc.fillna(0).iloc[:, -1]
    floating_to_percentage_month(HTO_bend_rate_per_month_mcc)

    SO_df['Week_number'] = pd.DataFrame(SO_perweek)
    SO_df['Month_number'] = pd.DataFrame(SO_permonth)
    SO_bend_rate_per_week = SO_df.groupby('Week_number')[['产品总个数', Type_of_defect]].sum()
    SO_bend_rate_per_week['bend_rate_per_week'] = (SO_bend_rate_per_week.iloc[:, 1] / SO_bend_rate_per_week.iloc[:, 0])
    SO_bend_rate_per_week = SO_bend_rate_per_week.reset_index()
    floating_to_percentage_week(SO_bend_rate_per_week)

    SO_bend_rate_per_week_mcc = SO_df.groupby(['Week_number', 'CDM机号'])[['产品总个数', Type_of_defect]].sum().unstack(
        fill_value=0).stack()
    SO_bend_rate_per_week_mcc['bend_rate_per_week'] = (
            SO_bend_rate_per_week_mcc.iloc[:, 1] / SO_bend_rate_per_week_mcc.iloc[:, 0])
    SO_bend_rate_per_week_mcc = SO_bend_rate_per_week_mcc.reset_index()
    SO_bend_rate_per_week_mcc.iloc[:, -1] = SO_bend_rate_per_week_mcc.fillna(0).iloc[:, -1]
    floating_to_percentage_week(SO_bend_rate_per_week_mcc)

    SO_bend_rate_per_month = SO_df.groupby('Month_number')[['产品总个数', Type_of_defect]].sum()
    SO_bend_rate_per_month['bend_rate_per_month'] = (
            SO_bend_rate_per_month.iloc[:, 1] / SO_bend_rate_per_month.iloc[:, 0])
    SO_bend_rate_per_month = SO_bend_rate_per_month.reset_index()
    floating_to_percentage_month(SO_bend_rate_per_month)

    SO_bend_rate_per_month_mcc = SO_df.groupby(['Month_number', 'CDM机号'])[['产品总个数', Type_of_defect]].sum().unstack(
        fill_value=0).stack()
    SO_bend_rate_per_month_mcc['bend_rate_per_month'] = (
            SO_bend_rate_per_month_mcc.iloc[:, 1] / SO_bend_rate_per_month_mcc.iloc[:, 0])
    SO_bend_rate_per_month_mcc = SO_bend_rate_per_month_mcc.reset_index()
    SO_bend_rate_per_month_mcc.iloc[:, -1] = SO_bend_rate_per_month_mcc.fillna(0).iloc[:, -1]
    floating_to_percentage_month(SO_bend_rate_per_month_mcc)

    CDM_SO_month = {}
    # try:
    #     CDM_SO_month['CDM1'] = (SO_bend_rate_per_month_mcc.groupby(SO_bend_rate_per_month_mcc.CDM机号).get_group(1)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_SO_month['CDM2'] = (SO_bend_rate_per_month_mcc.groupby(SO_bend_rate_per_month_mcc.CDM机号).get_group(2)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_SO_month['CDM3'] = (SO_bend_rate_per_month_mcc.groupby(SO_bend_rate_per_month_mcc.CDM机号).get_group(3)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_SO_month['CDM4'] = (SO_bend_rate_per_month_mcc.groupby(SO_bend_rate_per_month_mcc.CDM机号).get_group(4)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_SO_month['CDM5'] = (SO_bend_rate_per_month_mcc.groupby(SO_bend_rate_per_month_mcc.CDM机号).get_group(5)).reset_index(drop=True)
    # except KeyError:
    #     pass

    for i in range(1, 6):
        try:
            CDM_SO_month[f'CDM{i}'] = (
                (SO_bend_rate_per_month_mcc.groupby(SO_bend_rate_per_month_mcc.CDM机号).get_group(i)).reset_index(
                    drop=True))
        except KeyError:
            pass

    CDM_MPT_month = {}
    # try:
    #     CDM_MPT_month['CDM1'] = (MPT_bend_rate_per_month_mcc.groupby(MPT_bend_rate_per_month_mcc.CDM机号).get_group(1)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_MPT_month['CDM2'] = (MPT_bend_rate_per_month_mcc.groupby(MPT_bend_rate_per_month_mcc.CDM机号).get_group(2)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_MPT_month['CDM3'] = (MPT_bend_rate_per_month_mcc.groupby(MPT_bend_rate_per_month_mcc.CDM机号).get_group(3)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_MPT_month['CDM4'] = (MPT_bend_rate_per_month_mcc.groupby(MPT_bend_rate_per_month_mcc.CDM机号).get_group(4)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_MPT_month['CDM5'] = (MPT_bend_rate_per_month_mcc.groupby(MPT_bend_rate_per_month_mcc.CDM机号).get_group(5)).reset_index(drop=True)
    # except KeyError:
    #     pass

    for i in range(1, 6):
        try:
            CDM_MPT_month[f'CDM{i}'] = (
                (MPT_bend_rate_per_month_mcc.groupby(MPT_bend_rate_per_month_mcc.CDM机号).get_group(i)).reset_index(
                    drop=True))
        except KeyError:
            pass

    CDM_HTO_month = {}
    # try:
    #     CDM_HTO_month['CDM1'] = (HTO_bend_rate_per_month_mcc.groupby(HTO_bend_rate_per_month_mcc.CDM机号).get_group(1)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_HTO_month['CDM2'] = (HTO_bend_rate_per_month_mcc.groupby(HTO_bend_rate_per_month_mcc.CDM机号).get_group(2)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_HTO_month['CDM3'] = (HTO_bend_rate_per_month_mcc.groupby(HTO_bend_rate_per_month_mcc.CDM机号).get_group(3)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_HTO_month['CDM4'] = (HTO_bend_rate_per_month_mcc.groupby(HTO_bend_rate_per_month_mcc.CDM机号).get_group(4)).reset_index(drop=True)
    # except KeyError:
    #     pass
    # try:
    #     CDM_HTO_month['CDM5'] = (HTO_bend_rate_per_month_mcc.groupby(HTO_bend_rate_per_month_mcc.CDM机号).get_group(5)).reset_index(drop=True)
    # except KeyError:
    #     pass

    for i in range(1, 6):
        try:
            CDM_HTO_month[f'CDM{i}'] = (
                (HTO_bend_rate_per_month_mcc.groupby(HTO_bend_rate_per_month_mcc.CDM机号).get_group(i)).reset_index(
                    drop=True))
        except KeyError:
            pass

    CDM_SO_week = {}
    for i in range(1, 6):
        try:
            CDM_SO_week[f'CDM{i}'] = (
                (SO_bend_rate_per_week_mcc.groupby(SO_bend_rate_per_week_mcc.CDM机号).get_group(i)).reset_index(
                    drop=True))
        except KeyError:
            pass
    CDM_MPT_week = {}
    for i in range(1, 6):
        try:
            CDM_MPT_week[f'CDM{i}'] = (
                (MPT_bend_rate_per_week_mcc.groupby(MPT_bend_rate_per_week_mcc.CDM机号).get_group(i)).reset_index(
                    drop=True))
        except KeyError:
            pass
    CDM_HTO_week = {}
    for i in range(1, 6):
        try:
            CDM_HTO_week[f'CDM{i}'] = (
                (HTO_bend_rate_per_week_mcc.groupby(HTO_bend_rate_per_week_mcc.CDM机号).get_group(i)).reset_index(
                    drop=True))
        except KeyError:
            pass

    CDM_Total_week = {}
    for i in range(1, 6):
        try:
            CDM_Total_week[f'CDM{i}'] = (
                (Total_bend_rate_per_week_mcc.groupby(Total_bend_rate_per_week_mcc.CDM机号).get_group(i)).reset_index(
                    drop=True))
        except KeyError:
            pass

    CDM_numberoflot_week = (main_df.groupby(['CDM机号', 'Week_number'])[['生产批号']].count().unstack(fill_value=0).stack()). \
        reset_index()
    p = (CDM_numberoflot_week.groupby(CDM_numberoflot_week.CDM机号).get_group(5)).reset_index(drop=True)
    # print(p['生产批号'][2])
    CDM_numberoflot_month = (
        main_df.groupby(['CDM机号', 'Month_number'])[['生产批号']].count().unstack(fill_value=0).stack()). \
        reset_index()
    q = (CDM_numberoflot_month.groupby(CDM_numberoflot_month.CDM机号).get_group(5)).reset_index(drop=True)

    from Script.SortingandFiltering import Total_bend_rate_per_month_mcc, Total_bend_rate_per_month, \
        Total_bend_rate_per_week, SO_bend_rate_per_month, SO_bend_rate_per_week, MPT_bend_rate_per_month, \
        MPT_bend_rate_per_week, HTO_bend_rate_per_month, HTO_bend_rate_per_week, CDM_MPT_month, CDM_HTO_month, \
        CDM_SO_month, CDM_SO_week, CDM_MPT_week, CDM_HTO_week, CDM_Total_week, main_df, SO_permonth, SO_perweek, \
        MPT_permonth, MPT_perweek, HTO_permonth, HTO_perweek, CDM_numberoflot_week, CDM_numberoflot_month
    import pandas as pd
    from pandas import DataFrame
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook
    from Script.TemplateForWeeklyBendReport import bend_report_template, month_sorting_order, week_sorting_order, \
        new_week_sorting_order, new_month_sorting_order, template_date_list, sorting_month_template, \
        sorting_week_template, \
        lot_counting_template, lot_number_template
    from Script.Lot_counting import total_week_lot_list, total_week_lot_list_mmc, Lot_counting

    Constant_a = []

    wb = Workbook()
    ws = wb.active
    ws.title = "Week"

    bend_report_template(ws)

    for i in range(len(Total_bend_rate_per_month)):
        ws[get_column_letter(i + 3) + str(11)] = \
            (Total_bend_rate_per_month.loc[:, 'bend_rate_per_month'])[i]
        ws[get_column_letter(i + 3) + str(40)] = \
            (Total_bend_rate_per_month.loc[:, 'bend_rate_per_month'])[i]

    for i in range(len(SO_bend_rate_per_month)):
        ws[get_column_letter(i + 3 + sorting_month_template(SO_permonth)[i]) + str(41)] = \
            (SO_bend_rate_per_month.loc[:, 'bend_rate_per_month'])[i]

    for i in range(len(MPT_bend_rate_per_month)):
        ws[get_column_letter(i + 3 + sorting_month_template(MPT_permonth)[i]) + str(42)] = \
            (MPT_bend_rate_per_month.loc[:, 'bend_rate_per_month'])[i]

    for i in range(len(HTO_bend_rate_per_month)):
        ws[get_column_letter(i + 3 + sorting_month_template(HTO_permonth)[i]) + str(43)] = \
            (HTO_bend_rate_per_month.loc[:, 'bend_rate_per_month'])[i]

    Total_bend_rate_per_month_mcc = Total_bend_rate_per_month_mcc.sort_values((["CDM机号", 'Month_number']))
    for i in range(5):
        Constant_a = Total_bend_rate_per_month_mcc.iloc[(int(len(Total_bend_rate_per_month_mcc) / 5) * (i)) \
                                                        :(int(len(Total_bend_rate_per_month_mcc) / 5) * (1 + i)), :]
        for j in range(int(len(Total_bend_rate_per_month_mcc) / 5)):
            ws[get_column_letter(j + 3) + str(6 + i)] = \
                (list(Constant_a.loc[:, 'bend_rate_per_month']))[j]
            ws[get_column_letter(j + 3) + str(15 + i * 5)] = \
                (list(Constant_a.loc[:, 'bend_rate_per_month']))[j]

    Constant_b = [11, 40]
    for j in range(len(Constant_b)):
        for i in range(len(Total_bend_rate_per_week)):
            ws[get_column_letter(4 + len(template_date_list(main_df)) + i) + str(f'{Constant_b[j]}')] = \
                (list(Total_bend_rate_per_week.loc[:, 'bend_rate_per_week']))[i]

    # for i in range(len(Total_bend_rate_per_week)):
    #     ws[get_column_letter(6+new_week_sorting_order(Total_bend_rate_per_week)[i]) + str(40)] = \
    #         (list(Total_bend_rate_per_week.loc[:, 'bend_rate_per_week']))[i]

    # try:
    #     for i in range(len(CDM_SO_month['CDM1'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_SO_month['CDM1'])) + str(16)] = \
    #             (list(CDM_SO_month['CDM1'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_SO_month['CDM2'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_SO_month['CDM2'])) + str(21)] = \
    #             (list(CDM_SO_month['CDM2'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_SO_month['CDM3'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_SO_month['CDM3'])) + str(26)] = \
    #             (list(CDM_SO_month['CDM3'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_SO_month['CDM4'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_SO_month['CDM4'])) + str(31)] = \
    #             (list(CDM_SO_month['CDM4'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_SO_month['CDM5'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_SO_month['CDM5'])) + str(36)] = \
    #             (list(CDM_SO_month['CDM5'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    #
    # try:
    #     for i in range(len(CDM_MPT_month['CDM1'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_MPT_month['CDM1'])) + str(17)] = \
    #             (list(CDM_MPT_month['CDM1'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_MPT_month['CDM2'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_MPT_month['CDM2'])) + str(22)] = \
    #             (list(CDM_MPT_month['CDM2'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_MPT_month['CDM3'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_MPT_month['CDM3'])) + str(27)] = \
    #             (list(CDM_MPT_month['CDM3'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_MPT_month['CDM4'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_MPT_month['CDM4'])) + str(32)] = \
    #             (list(CDM_MPT_month['CDM4'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_MPT_month['CDM5'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_MPT_month['CDM5'])) + str(37)] = \
    #             (list(CDM_MPT_month['CDM5'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass

    # try:
    #     for i in range(len(CDM_HTO_month['CDM1'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_HTO_month['CDM1'])) + str(18)] = \
    #             (list(CDM_HTO_month['CDM1'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_HTO_month['CDM2'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_HTO_month['CDM2'])) + str(23)] = \
    #             (list(CDM_HTO_month['CDM2'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_HTO_month['CDM3'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_HTO_month['CDM3'])) + str(28)] = \
    #             (list(CDM_HTO_month['CDM3'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_HTO_month['CDM4'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_HTO_month['CDM4'])) + str(33)] = \
    #             (list(CDM_HTO_month['CDM4'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass
    # try:
    #     for i in range(len(CDM_HTO_month['CDM5'])):
    #         ws[get_column_letter(i + 2 + month_sorting_order(CDM_HTO_month['CDM5'])) + str(38)] = \
    #             (list(CDM_HTO_month['CDM5'].loc[:, 'bend_rate_per_month']))[i]
    # except KeyError:
    #     pass

    # for i in range(len(CDM_SO_week['CDM2'])):
    #     ws[get_column_letter(i + 15 + week_sorting_order(CDM_SO_week['CDM2'])) + str(21)] = \
    #         (list(CDM_SO_week['CDM2'].loc[:, 'bend_rate_per_week']))[i]

    # for i in range(1, 6):
    #     try:
    #         for x in range(len(CDM_SO_week[f'CDM{i}'])):
    #             ws[get_column_letter(x + 15 + week_sorting_order(CDM_SO_week[f'CDM{i}'])) + str(16+5*(i-1))] = \
    #                 (list(CDM_SO_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
    #     except KeyError:
    #         pass
    # for i in range(1, 6):
    #     try:
    #         for x in range(len(CDM_MPT_week[f'CDM{i}'])):
    #             ws[get_column_letter(x + 15 + week_sorting_order(CDM_MPT_week[f'CDM{i}'])) + str(17+5*(i-1))] = \
    #                 (list(CDM_MPT_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
    #     except KeyError:
    #         pass
    # for i in range(1, 6):
    #     try:
    #         for x in range(len(CDM_HTO_week[f'CDM{i}'])):
    #             ws[get_column_letter(x + 15 + week_sorting_order(CDM_HTO_week[f'CDM{i}'])) + str(18+5*(i-1))] = \
    #                 (list(CDM_HTO_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
    #     except KeyError:
    #         pass

    for i in range(len(SO_bend_rate_per_week)):
        ws[get_column_letter(4 + len(template_date_list(main_df)) + i + sorting_week_template(SO_perweek)[i]) + str(
            41)] = \
            (list(SO_bend_rate_per_week.loc[:, 'bend_rate_per_week']))[i]

    for i in range(len(MPT_bend_rate_per_week)):
        ws[get_column_letter(4 + len(template_date_list(main_df)) + i + sorting_week_template(MPT_perweek)[i]) + str(
            42)] = \
            (list(MPT_bend_rate_per_week.loc[:, 'bend_rate_per_week']))[i]

    for i in range(len(HTO_bend_rate_per_week)):
        ws[get_column_letter(4 + len(template_date_list(main_df)) + i + sorting_week_template(HTO_perweek)[i]) + str(
            43)] = \
            (list(HTO_bend_rate_per_week.loc[:, 'bend_rate_per_week']))[i]

    for i in range(1, 6):
        try:
            for x in range(len(CDM_Total_week[f'CDM{i}'])):
                ws[get_column_letter(4 + len(template_date_list(main_df)) + x) + str(15 + 5 * (i - 1))] = \
                    (list(CDM_Total_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
                ws[get_column_letter(4 + len(template_date_list(main_df)) + x) + str(6 + 1 * (i - 1))] = \
                    (list(CDM_Total_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
        except KeyError:
            pass

    for i in range(1, 6):
        try:
            for x in range(len(CDM_HTO_week[f'CDM{i}'])):
                ws[get_column_letter(4 + len(template_date_list(main_df)) + x + sorting_week_template(HTO_perweek)[x]) + \
                   str(18 + 5 * (i - 1))] = (list(CDM_HTO_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
        except KeyError:
            pass
    for i in range(1, 6):
        try:
            for x in range(len(CDM_HTO_month[f'CDM{i}'])):
                ws[get_column_letter(x + 3 + sorting_month_template(HTO_permonth)[x]) + str(18 + 5 * (i - 1))] = \
                    (list(CDM_HTO_month[f'CDM{i}'].loc[:, 'bend_rate_per_month']))[x]
        except KeyError:
            pass
    for i in range(1, 6):
        try:
            for x in range(len(CDM_MPT_week[f'CDM{i}'])):
                ws[get_column_letter(4 + len(template_date_list(main_df)) + x + sorting_week_template(MPT_perweek)[x]) + \
                   str(17 + 5 * (i - 1))] = (list(CDM_MPT_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
        except KeyError:
            pass
    for i in range(1, 6):
        try:
            for x in range(len(CDM_MPT_month[f'CDM{i}'])):
                ws[get_column_letter(x + 3 + sorting_month_template(MPT_permonth)[x]) + str(17 + 5 * (i - 1))] = \
                    (list(CDM_MPT_month[f'CDM{i}'].loc[:, 'bend_rate_per_month']))[x]
        except KeyError:
            pass
    for i in range(1, 6):
        try:
            for x in range(len(CDM_SO_week[f'CDM{i}'])):
                ws[get_column_letter(4 + len(template_date_list(main_df)) + x + sorting_week_template(SO_perweek)[x]) + \
                   str(16 + 5 * (i - 1))] = (list(CDM_SO_week[f'CDM{i}'].loc[:, 'bend_rate_per_week']))[x]
        except KeyError:
            pass
    for i in range(1, 6):
        try:
            for x in range(len(CDM_SO_month[f'CDM{i}'])):
                ws[get_column_letter(x + 3 + sorting_month_template(SO_permonth)[x]) + str(16 + 5 * (i - 1))] = \
                    (list(CDM_SO_month[f'CDM{i}'].loc[:, 'bend_rate_per_month']))[x]
        except KeyError:
            pass

    for row in range(6):
        for cell in range(len(Total_bend_rate_per_month)):
            (ws[get_column_letter(cell + 3) + str(6 + row)]) \
                .number_format = "0.00%"
        for cell in range(len(Total_bend_rate_per_week)):
            (ws[get_column_letter(4 + len(template_date_list(main_df)) + cell) + str(6 + row)]) \
                .number_format = "0.00%"
    for row in range(29):
        for cell in range(len(Total_bend_rate_per_month)):
            (ws[get_column_letter(cell + 3) + str(15 + row)]) \
                .number_format = "0.00%"
        for cell in range(len(Total_bend_rate_per_week)):
            (ws[get_column_letter(4 + len(template_date_list(main_df)) + cell) + str(15 + row)]) \
                .number_format = "0.00%"

    new_worksheet = wb.create_sheet('lot_counting')
    ws2 = wb['lot_counting']
    lot_counting_template(ws2)

    for i in range(3):
        SO_value, SO_machine = Lot_counting('SO')
        HTO_value, HTO_machine = Lot_counting('HTO')
        MPT_value, MPT_machine = Lot_counting('MPT')
        ws2['C' + str(4 + i)] = len(total_week_lot_list[i])
        ws2['E' + str(4 + i)] = SO_value[i]
        ws2['G' + str(4 + i)] = MPT_value[i]
        ws2['I' + str(4 + i)] = HTO_value[i]
        for x in range(5):
            ws2['C' + str(10 + x * 6 + 1 * i)] = len(total_week_lot_list_mmc[5 * i + x])
            ws2['E' + str(10 + x * 6 + 1 * i)] = SO_machine[5 * i + x]
            ws2['G' + str(10 + x * 6 + 1 * i)] = MPT_machine[5 * i + x]
            ws2['I' + str(10 + x * 6 + 1 * i)] = HTO_machine[5 * i + x]

    for i in range(1, 6):
        try:
            for x in range(len(CDM_Total_week[f'CDM{i}'][-3:])):
                ws2['D' + str(12 - x * 1 + (i - 1) * 6)] = \
                    (CDM_Total_week[f'CDM{i}'][-3:].iloc[x, -1])
        except KeyError:
            pass
    Three_weeks = sorted(list(main_df['Week_number'].drop_duplicates()))[-3:]
    for i in range(1, 6):
        for x in range(len(Three_weeks)):
            try:
                ws2['J' + str(12 + (i - 1) * 6 - 1 * x)] = (
                CDM_HTO_week[f'CDM{i}'][CDM_HTO_week[f'CDM{i}'].Week_number == \
                                        Three_weeks[x]].iloc[0, -1])
            except IndexError:
                pass
            except KeyError:
                pass
        for x in range(len(Three_weeks)):
            try:
                ws2['F' + str(12 + (i - 1) * 6 - 1 * x)] = (CDM_SO_week[f'CDM{i}'][CDM_SO_week[f'CDM{i}'].Week_number == \
                                                                                   Three_weeks[x]].iloc[0, -1])
            except IndexError:
                pass
            except KeyError:
                pass
        for x in range(len(Three_weeks)):
            try:
                ws2['H' + str(12 + (i - 1) * 6 - 1 * x)] = (
                CDM_MPT_week[f'CDM{i}'][CDM_MPT_week[f'CDM{i}'].Week_number == \
                                        Three_weeks[x]].iloc[0, -1])
            except IndexError:
                pass
            except KeyError:
                pass

    for x in range(len(Three_weeks)):
        try:
            ws2['J' + str(6 - 1 * x)] = (HTO_bend_rate_per_week[HTO_bend_rate_per_week.Week_number == \
                                                                Three_weeks[x]].iloc[0, -1])
        except IndexError:
            pass
        except KeyError:
            pass
    for x in range(len(Three_weeks)):
        try:
            ws2['F' + str(6 - 1 * x)] = (SO_bend_rate_per_week[SO_bend_rate_per_week.Week_number == \
                                                               Three_weeks[x]].iloc[0, -1])
        except IndexError:
            pass
        except KeyError:
            pass
    for x in range(len(Three_weeks)):
        try:
            ws2['H' + str(6 - 1 * x)] = (MPT_bend_rate_per_week[MPT_bend_rate_per_week.Week_number == \
                                                                Three_weeks[x]].iloc[0, -1])
        except IndexError:
            pass
        except KeyError:
            pass
    for x in range(len(Three_weeks)):
        try:
            ws2['D' + str(6 - 1 * x)] = (Total_bend_rate_per_week[Total_bend_rate_per_week.Week_number == \
                                                                  Three_weeks[x]].iloc[0, -1])
        except IndexError:
            pass
        except KeyError:
            pass

    # (CDM_Total_week['CDM2'][-3:].iloc[2, -1])
    #
    # for i in range (len(All_HTO_lots)):
    #     ws['C' + str(i + 2)] = All_HTO_lots[i]
    #
    # for i in range (len(All_SO_lots)):
    #     ws['D' + str(i + 2)] = All_SO_lots[i]
    #
    # for i in range (len(All_Other_lots)):
    #     ws['E' + str(i + 2)] = All_Other_lots[i]
    #

    new_worksheet = wb.create_sheet('lot_number')
    ws3 = wb['lot_number']
    lot_number_template(ws3)
    for x in range(1, 6):
        for i in range(Total_bend_rate_per_week.shape[0]):
            ws3[get_column_letter(11 + i) + str(5 + x)] = (
            CDM_numberoflot_week.groupby(CDM_numberoflot_week.CDM机号).get_group(x). \
                reset_index(drop=True)['生产批号'][i])
        for j in range(Total_bend_rate_per_month.shape[0]):
            ws3[get_column_letter(3 + j) + str(5 + x)] = (CDM_numberoflot_month.groupby(CDM_numberoflot_month.CDM机号). \
                get_group(x).reset_index(drop=True)['生产批号'][j])

    wb.save('Weektestsummary.xlsx')



def input():
    file1 = filedialog.askopenfile()
    file_list = file1
    return file_list
def input2():
    file2 = filedialog.asksaveasfile(mode="w", defaultextension=".tif")
    label = Label(text=file2).pack()



w = Tk()
w.geometry("500x500")
w.title("FLOOD_MAPPER")
h = Label(text = "S1A FLOOD MAPPER", bg = "yellow", fg = "black", height = "3", width = "500")
h.pack()
myButton1 = Checkbutton(w, text="contamination",)
myButton1.pack()
i1 = Label(text = "Input*")
i1.place(x=10, y=70)
i1b = Button(w, text = "Select File", command =input)
i1b.place(x=250, y=70)
i2 = Label(text = "Intermediate Product*")
i2.place(x=10, y=140)
i2b = Button(w, text = "Save as", command =input2)
i2b.place(x=250, y=140)
if i1b:
    button = Button(w, text="Generate Map", bg = "red", fg = "black", height = "2", width="30", command = excel_input('Co + Adhesive Co', input()))
    button.place(x=150, y=400)
w.mainloop()