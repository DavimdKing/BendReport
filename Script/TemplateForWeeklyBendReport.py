from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from Script.SortingandFiltering import Total_bend_rate_per_week, main_df
import re
import pandas as pd

def date_format_tranform(time):
    new_format = time.strftime('%b %y')
    return new_format


def template_date_list(df):
    df['month_year'] = df['CDM生产时间'].apply(date_format_tranform)
    first_date = (df['CDM生产时间'][0])
    last_date = (df['CDM生产时间'].iloc[-1])
    series_of_month = (df['month_year']).drop_duplicates()
    list_of_month = series_of_month.tolist()
    return list_of_month




def bend_report_template(ws):
    All_Month = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06',
                 'July': '07',
                 'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12'}
    Product_type = ['Total', 'SO', 'MPT', 'HTO', '']


    for i in range(5):
        ws['B' + str(i + 6)] = 'CDM' + str(i + 1)
    ws['B11'] = 'Average'

    for index in range(len(template_date_list(main_df))):
        ws[get_column_letter(index + 3) + str(5)] = template_date_list(main_df)[index]
    for index in range(Total_bend_rate_per_week.shape[0]):
        ws[get_column_letter(index + 4 + len(template_date_list(main_df))) + str(5)] = 'W' + str(
            Total_bend_rate_per_week['Week_number'][index][4:6])

    for index in range(len(template_date_list(main_df))):
        ws[get_column_letter(index + 3) + str(14)] = template_date_list(main_df)[index]
    for index in range(Total_bend_rate_per_week.shape[0]):
        ws[get_column_letter(index + 4 + len(template_date_list(main_df))) + str(14)] = 'W' + str(
            Total_bend_rate_per_week['Week_number'][index][4:6])

    for i in range(5):
        ws['A' + str(i * 5 + 15)] = 'CDM' + str(i + 1)
    ws['A40'] = 'Average'

    for i in range(6):
        for index in range(len(Product_type)):
            ws['B' + str(index + 15 + 5 * i)] = Product_type[index]

    # wb.save('../Output/Week38summary.xlsx')

def month_sorting_order(i):
    All_Month = {'01January': '01', '02February': '02', '03March': '03', '04April': '04', '05May': '05', '06June': '06',
                 '07July': '07', '08August': '08',
                 '09September': '09', '10October': '10', '11November': '11', '12December': '12'}
    y = i.loc[:,'Month_number'][0]
    z = 1
    for x in All_Month:
        if x == y:
            break
        z = z + 1
    return z

def week_sorting_order(i):
    ALL_Week = []
    for x in range(10,53):
        ALL_Week.append("Week" + str(x))
    z = 1
    y = i.loc[:, 'Week_number'][0]
    for x in ALL_Week:
        if x == y:
            break
        z = z + 1
    return z

def new_week_sorting_order(i):
    All_weeks = []
    y = i.loc[:, 'Week_number']
    for x in y:
        All_weeks.append(int(x.split(sep='k')[1]))
    return  All_weeks

def new_month_sorting_order(i):
    All_months = []
    y = i.loc[:, 'Month_number']
    for x in y:
        All_months.append(int((re.findall('\d+', x))[0]))
    return All_months

def sorting_month_template(customer_permonth):
    Constant_c = []
    Constant_d = 0
    Constant_e = []
    if len(list(main_df['Month_number'].drop_duplicates())) == len(list(dict.fromkeys(customer_permonth))):
        for i in range(len(list(main_df['Month_number'].drop_duplicates()))):
            Constant_e.append(0)
    else:
        for i in range(len(list(main_df['Month_number'].drop_duplicates()))):
            if (sorted(list(main_df['Month_number'].drop_duplicates())))[i] in list(dict.fromkeys(customer_permonth)):
                Constant_e.append(Constant_d)
            else:
                Constant_d = Constant_d + 1
    return Constant_e

def sorting_week_template(customer_perweek):
    Constant_c = []
    Constant_d = 0
    Constant_e = []
    if len(list(main_df['Week_number'].drop_duplicates())) == len(list(dict.fromkeys(customer_perweek))):
        for i in range(len(list(main_df['Week_number'].drop_duplicates()))):
            Constant_e.append(0)
    else:
        for i in range(len(list(main_df['Week_number'].drop_duplicates()))):
            if (sorted(list(main_df['Week_number'].drop_duplicates())))[i] in list(dict.fromkeys(customer_perweek)):
                Constant_e.append(Constant_d)
            else:
                Constant_d = Constant_d + 1
    return Constant_e

def lot_counting_template(ws2):
    Three_weeks = sorted(list(main_df['Week_number'].drop_duplicates()))[-3:]
    Three_weeks.reverse()
    CDM_list = ['All', 'CDM1', 'CDM2', 'CDM3', 'CDM4', 'CDM5']
    Column_name = ['Total', 'So', 'MPT', 'HTO']
    sub_column_name = ['Lot:N', 'Rate']
    for x in range(6):
        for i in range(len(Three_weeks)):
            ws2['B' + str(4+i+x*6)] = Three_weeks[i]
        ws2['B' + str(2+x*6)] = CDM_list[x]
        var1 = 2+x*6
        var2 = 3+x*6
        ws2.merge_cells(f"B{var1}:B{var2}")

        for i in range(len(Column_name)):
            ws2[get_column_letter(3+i*2) + str(2+x*6)] = Column_name[i]
            var3 = get_column_letter(3+i*2) + str(2+x*6)
            var4 = get_column_letter(4+i*2) + str(2+x*6)
            ws2.merge_cells(f"{var3}:{var4}")
            ws2[get_column_letter(3 + i * 2) + str(3 + x * 6)] = sub_column_name[0]
            ws2[get_column_letter(4 + i * 2) + str(3 + x * 6)] = sub_column_name[1]


def lot_number_template(ws3):
    All_Month = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06',
                 'July': '07',
                 'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12'}
    Product_type = ['Total', 'SO', 'MPT', 'HTO', '']

    for i in range(5):
        ws3['B' + str(i + 6)] = 'CDM' + str(i + 1)

    for index in range(len(template_date_list(main_df))):
        ws3[get_column_letter(index + 3) + str(5)] = template_date_list(main_df)[index]
    for index in range(Total_bend_rate_per_week.shape[0]):
        ws3[get_column_letter(index + 4 + len(template_date_list(main_df))) + str(5)] = 'W' + str(
            Total_bend_rate_per_week['Week_number'][index][4:6])

