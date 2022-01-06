import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
import glob
import ntpath
from append_data_in_excel import append_df_to_excel

# wb= load_workbook("C:/Users/wheng/Downloads/VEM data/COPY/VEM 2 copy/A211205-A-3A.CSV")



folder_file = []

for filename in glob.glob("C:/Users/wheng/Downloads/VEM data/New_wk35_36/*"):
    folder_file.append(filename)

for i in range(8):

    VEM_file = []
    for filename in glob.glob(folder_file[i] + "/*.CSV"):

        VEM_file.append(filename)

    excel_df = {}
    sheet_list = []
    coordinate_no_list = []
    Bad_cycletime_list = []
    Good_cycletime_list = []
    vacuum_time_list = []
    exposure_time_list = []
    lot_name_list = []
    start_time_list = []
    last_time_list = []
    for i in range(len(VEM_file)):



        main_df = pd.read_csv(VEM_file[i], header=5,
                                  encoding="ISO-8859-1")

    # for filename in glob.glob("C:/Users/wheng/Downloads/VEM data/COPY/VEM 2 copy/*.CSV"):
    #     folder_file.append(filename)
    #
    # for i in range(len(folder_file)):
    #     main_df = pd.read_csv(folder_file[i], header=5,
    #                           encoding="ISO-8859-1")
        main_df = (main_df.iloc[3:, :].reset_index(drop=True))
        start_time = (main_df[":DEV_COMMENT"].iloc[0])
        start_time_list.append(start_time)
        last_time = (main_df[":DEV_COMMENT"].iloc[-1])
        last_time_list.append(last_time)
        lot_name = ntpath.basename(VEM_file[i])
        lot_name_list.append(lot_name)

        sheet_number = int(main_df.iloc[-1, 3])
        sheet_list.append(sheet_number)

        coordinate_no = len(main_df.iloc[:, 2])
        coordinate_no_list.append(coordinate_no)

        Bad_cycletime = main_df['Cycle Period'].astype(int)
        Bad_cycletime = int(Bad_cycletime.where(Bad_cycletime <= 110).sum(axis=0))
        Bad_cycletime_list.append(Bad_cycletime)

        Good_cycletime = main_df['Cycle Period'].astype(int)
        Good_cycletime = int(Good_cycletime.where(Good_cycletime >= 110).sum(axis=0))
        Good_cycletime_list.append(Good_cycletime)

        vacuum_time = (main_df['1st Time'].astype(int).sum(axis=0)) / 10
        vacuum_time_list.append(vacuum_time)

        exposure_time = (main_df['A Exp Time'].astype(int).sum(axis=0))
        exposure_time_list.append(exposure_time)

    excel_df['lot_number'] = lot_name_list
    excel_df['開始時間'] = start_time_list
    excel_df['結束時間'] = last_time_list
    excel_df['張數'] = sheet_list
    excel_df['定位次數'] = coordinate_no_list
    excel_df['NG張總處理時間(seconds)'] = Bad_cycletime_list
    excel_df['非NG張總生產時間(seconds)'] = Good_cycletime_list
    excel_df['抽真空#1時間'] = vacuum_time_list
    excel_df['曝光总时间（秒）'] = exposure_time_list
    excel_df = (pd.DataFrame(excel_df))

    append_df_to_excel("NewResultofSpecialmission.xlsx", excel_df, sheet_name='VEM_ALL')

# wb = load_workbook("ResultofSpecialmission.xlsx")
# writer = pd.ExcelWriter('ResultofSpecialmission.xlsx', engine='openpyxl')
# writer.book = wb
# writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
# excel_df.to_excel(writer, 'VEM')
# writer.save()
# writer.close()

# excel_df.to_excel('ResultofSpecialmission.xlsx', mode='a')
